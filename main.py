import tkinter as tk
from tkinter import filedialog, Entry, Label, Button
from openpyxl import load_workbook
import sqlite3
import os
from openpyxl.styles import PatternFill
from gui import root
from renaming_of_professions import renaming_of_professions
from loguru import logger

# Определяем название базы данных и таблицы
file_database = "database.db"
table_name = "my_table"

# Создаем текст (метка)
label = tk.Label(root, text="Нажмите на кнопку ниже, чтобы выбрать файл:")
label.pack(pady=20)


# Функция для выбора файла
def select_file():
    file_path = filedialog.askopenfilename()
    if file_path:
        logger.info(f"Выбран файл: {file_path}")


# Функция открытия файла Excel
def opening_a_file():
    filename = filedialog.askopenfilename(filetypes=[('Excel Files', '*.xlsx')])
    return filename


# Функция обработки данных из Excel и вставки в базу данных SQLite
def parsing_document(min_row, max_row, column):
    filename = opening_a_file()  # Открываем выбор файла Excel для чтения данных
    workbook = load_workbook(filename=filename)  # Загружаем выбранный файл Excel
    sheet = workbook.active

    # Удаляем файл базы данных, если он существует
    if os.path.exists(file_database):
        os.remove(file_database)

    # Создаем соединение с базой данных
    conn = sqlite3.connect(file_database)
    cursor = conn.cursor()
    # Создаем таблицу в базе данных, если она еще не существует
    cursor.execute(f"CREATE TABLE IF NOT EXISTS {table_name} (table_column_1 TEXT)")

    # Считываем данные из колонки и вставляем их в базу данных
    for row in sheet.iter_rows(min_row=int(min_row), max_row=int(max_row), values_only=True):
        table_column_1 = str(row[int(column)])  # Преобразуем значение в строку
        cursor.execute(f"SELECT * FROM {table_name} WHERE table_column_1 = ?", (table_column_1,))
        existing_row = cursor.fetchone()
        if existing_row is None:
            cursor.execute(f"INSERT INTO {table_name} VALUES (?)", (table_column_1,))

    # Удаляем повторы по табельному номеру
    cursor.execute(
        f"DELETE FROM {table_name} WHERE rowid NOT IN (SELECT min(rowid) FROM {table_name} GROUP BY table_column_1)"
    )
    conn.commit()
    conn.close()


# Функция обработки кнопки "Готово"
def handle_done_button(entry1, entry2, entry3):
    logger.info("Данные введены:", entry1.get(), entry2.get(), entry3.get())
    parsing_document(entry1.get(), entry2.get(), entry3.get())


# Функция для графического ввода данных
def input_function():

    # Очистка содержимого окна и замена главного окна на новое окно
    for widget in root.winfo_children():
        widget.pack_forget()

    root.geometry('400x400')  # Устанавливаем размер окна

    # Создаем три поля ввода
    entry1 = Entry(root)
    entry2 = Entry(root)
    entry3 = Entry(root)

    # Добавляем метки к полям ввода
    Label(root, text="Минимальная строка:").grid(row=0, column=0, padx=5, pady=5)
    Label(root, text="Максимальная строка:").grid(row=2, column=0, padx=5, pady=5)
    Label(root, text="Столбец:").grid(row=4, column=0, padx=5, pady=5)

    # Размещаем поля ввода
    entry1.grid(row=1, column=0, padx=5, pady=5)
    entry2.grid(row=3, column=0, padx=5, pady=5)
    entry3.grid(row=5, column=0, padx=5, pady=5)

    # Увеличиваем ширину полей ввода
    entry1.config(width=70)
    entry2.config(width=70)
    entry3.config(width=70)

    # Создаем кнопку "Готово" и добавляем обработчик событий
    Button(root, text="Готово", command=lambda: handle_done_button(entry1, entry2, entry3)).grid(row=6, column=0, pady=10)



def find_and_highlight_duplicates():
    filename = opening_a_file()  # Открываем выбор файла Excel для чтения данных
    workbook = load_workbook(filename=filename)  # Загружаем выбранный файл Excel
    sheet = workbook.active
    
    # Создаем множество для хранения уникальных значений из файла и базы данных
    unique_values = set()
    duplicates = set()
    
    # Задаем стиль подсветки для дубликатов
    fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    
    # Создаем соединение с базой данных
    conn = sqlite3.connect(file_database)
    cursor = conn.cursor()
    
    # Извлекаем все значения из базы данных, приводим их к строкам
    cursor.execute(f"SELECT table_column_1 FROM {table_name}")
    db_values = set([str(row[0]) for row in cursor.fetchall()])
    logger.info(f"Данные из базы данных: {db_values}")
    
    # Добавляем значения из базы данных в уникальные значения
    unique_values.update(db_values)

    # Поиск дубликатов в файле и проверка их наличия в базе данных
    for row in range(7, 210):  # Проходим по строкам
        cell_value = sheet.cell(row=row, column=6).value  # Извлекаем значение из столбца 6 (индексация с 1)
        logger.info(f"Строка {row}, значение в ячейке: {cell_value}")
        
        # Приводим значение из ячейки к строке для корректного сравнения
        if cell_value is None:
            continue
        cell_value_str = str(cell_value)
        
        # Проверяем, есть ли значение в уникальных значениях (включая данные из базы)
        if cell_value_str in unique_values:
            duplicates.add(cell_value_str)
            logger.info(f"Найден дубликат: {cell_value_str}")
        else:
            unique_values.add(cell_value_str)  # Добавляем в уникальные значения

    # Подсветка дубликатов
    for row in range(7, 210):
        cell_value = sheet.cell(row=row, column=6).value
        cell_value_str = str(cell_value)
        if cell_value_str in duplicates:
            sheet.cell(row=row, column=6).fill = fill
            logger.info(f"Ячейка в строке {row} помечена как дубликат")

    # Сохранение изменений в файле
    workbook.save(filename)
    workbook.close()

    # Закрываем соединение с базой данных
    conn.close()
    logger.info('Работа окончена')




# Создаем кнопку "Выбрать файл"
button = tk.Button(root, text="Выбрать файл", command=select_file, width=48, height=1)
button.pack(pady=2)  # Расстояние между кнопками

# Создаем кнопку "Парсинг файла"
button1 = tk.Button(root, text="Парсинг файла", command=input_function, width=48, height=1)
button1.pack(pady=2)  # Расстояние между кнопками

# Создаем кнопку "Переименование профессий с EСXELL файла"
button3 = tk.Button(root, text="Переименование профессий с EСXELL файла", command=renaming_of_professions, width=48,
                    height=1)
button3.pack(pady=2)  # Расстояние между кнопками


# Создаем кнопку "Переименование профессий с EСXELL файла"
button4 = tk.Button(root, text="Сравнение данных и пометка цветом", command=find_and_highlight_duplicates, width=48,
                    height=1)
button4.pack(pady=2)  # Расстояние между кнопками


# Запуск главного цикла окна
root.mainloop()
