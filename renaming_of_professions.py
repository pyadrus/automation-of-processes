from tkinter import filedialog, Entry, Label, Button

from openpyxl import load_workbook

from gui import root


def parsing_document_1(min_row, max_row, column):
    """
    Функция обработки ввода данных
    min_row = 6 # Начальная строка
    max_row = 16 # Конечная строка
    column = 3 # Начальная колонка (счет начинается с 1)
    """

    filename = filedialog.askopenfilename(filetypes=[('Excel Files', '*.xlsx')])  # Выбор файла
    print(f"Выбран файл: {filename}")

    workbook = load_workbook(filename=filename)  # Загружаем выбранный файл Excel
    sheet = workbook.active

    # Считываем данные из колонки и заменяем их при необходимости
    for row in range(min_row, max_row + 1):
        cell = sheet.cell(row=row, column=column)  # Получаем ячейку
        table_column_1 = str(cell.value)  # Преобразуем значение в строку

        print(table_column_1)  # Выводим значение

        if table_column_1 == "нач. сектора (расчётного)":
            cell.value = "Начальник сектора расчётного"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "нач.уч-ка":
            cell.value = "Начальник участка"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "гл.бухгалтер":
            cell.value = "Главный бухгалтер"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "вед. экономист":
            cell.value = "Ведущий экономист"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "нач.отдела":
            cell.value = "Начальник отдела"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "нач. сектора":
            cell.value = "Начальник сектора"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "зам.гл.бух":
            cell.value = "Заместитель главного бухгалтера"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "зам. директора по экономике":
            cell.value = "Заместитель директора по экономике"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "бухгалтер 1 кат.(расчётного сектора)":
            cell.value = "Бухгалтер 1 категории расчётного сектора"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "уборщик служеб.помещений":
            cell.value = "Уборщик служебных помещений"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "мех.уч-ка":
            cell.value = "Механик участка"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "нач.смены":
            cell.value = "Начальник смены"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "гл.инженер":
            cell.value = "Главный инженер"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "зам. гл. инженера (по ПАЗ)":
            cell.value = "Заместитель главного инженера по противоаварийной защите"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "зам. директора (по ОТ и ТБ)":
            cell.value = "Заместитель директора по охране труда и ТБ"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "маш. комп.уст.":
            cell.value = "Машинист компрессорных установок"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "электрос.(слес.)деж.и по рем.оборудования":
            cell.value = "Электрослесарь (слесарь) дежурный и по ремонту оборудования"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "зав. гостиницы":
            cell.value = "Заведующий гостиницы"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "эл.слес.подз.":
            cell.value = "Электрослесарь подземный"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "зам.нач. уч-ка":
            cell.value = "Заместитель начальника участка"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "эл.монтажник по осв":
            cell.value = "Электромонтажник по освещению и осветительным сетям"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "лаборант хим.анализа":
            cell.value = "Лаборант химического анализа"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "зав. лаб. (углехимической)":
            cell.value = "Заведующий лабораторией углехимической"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "гл. инженер (ОФ)":
            cell.value = "Главный инженер обогатительной фабрики"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "зам. гл. инженера (ОФ)":
            cell.value = "Заместитель главного инженера обогатительной фабрики"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "зав.складом":
            cell.value = "Заведующий складом"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "и.о. нач. уч-ка":
            cell.value = "и.о. начальника участка"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "врач (кардиолог) высш. кат.-зав. здравпунктом":
            cell.value = "Врач-кардиолог высшей категории-заведующий здравпунктом"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "зав.общежитием":
            cell.value = "Заведующий общежитием"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "нач. производства (основного)":
            cell.value = "Начальник производства основного"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "маш.экск.":
            cell.value = "Машинист экскаватора"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "маш.кр.авт":
            cell.value = "Машинист крана автомобильного"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "мастер службы (техн. связи)":
            cell.value = "Мастер службы технологической связи"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "зав. складом (ВМ)":
            cell.value = "Заведующий складом взрывчатых материалов"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "мастер уч-ка":
            cell.value = "Мастер участка"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "гл.механик":
            cell.value = "Главный механик"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "гл. технолог":
            cell.value = "Главный технолог"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "и.о. зам.директора по экономике":
            cell.value = "и.о. заместителя директора по экономике"# Заменяем неправильную профессию на правильную
        elif table_column_1 == "ст. инспектор по контролю за исполнением поручений":
            cell.value = "Старший инспектор по контролю за исполнением поручений"# Заменяем неправильную профессию на правильную

        elif table_column_1 == "вед. инженер по ОТ (ТБ, учёту и анализу травматизма)":
            cell.value = "Ведущий инженер по охране труда ТБ, учёту и анализу травматизма"# Заменяем неправильную профессию на правильную

        elif table_column_1 == "вр.и.о. директора":
            cell.value = "врио директора"# Заменяем неправильную профессию на правильную

        elif table_column_1 == "и.о. нач. отдела":
            cell.value = "и.о. начальника отдела"# Заменяем неправильную профессию на правильную

        elif table_column_1 == "и.о. зам.нач. уч-ка":
            cell.value = "и.о. заместителя начальника участка"# Заменяем неправильную профессию на правильную

    try:
        # Сохраняем изменения в том же файле
        workbook.save(filename)
        print(f"Изменения сохранены в файле: {filename}")
    except PermissionError:
        print(f"Файл {filename} открыт в другой программе")

# Функция обработки кнопки "Готово"
def handle_done_button(entry1, entry2, entry3):
    print("Данные введены:", entry1.get(), entry2.get(), entry3.get())
    parsing_document_1(int(entry1.get()), int(entry2.get()), int(entry3.get()))

# Функция для графического ввода данных
def input_function_1():
    # Создаем главное окно
    root.title("Выбор файла")  # Заголовок окна

    root.geometry("400x400")  # Размер окна

    # Очистка содержимого окна и замена главного окна на новое окно
    for widget in root.winfo_children():
        widget.pack_forget()

    root.geometry('400x400')  # Устанавливаем размер окна

    # Создаем три поля ввода
    entry1 = Entry(root)
    entry2 = Entry(root)
    entry3 = Entry(root)

    # Добавляем метки к полям ввода
    Label(root, text="Минимальная строка:").grid(row=0, column=0, padx=5, pady=5)
    Label(root, text="Максимальная строка:").grid(row=2, column=0, padx=5, pady=5)
    Label(root, text="Столбец (счет начинается с 1) :").grid(row=4, column=0, padx=5, pady=5)

    # Размещаем поля ввода
    entry1.grid(row=1, column=0, padx=5, pady=5)
    entry2.grid(row=3, column=0, padx=5, pady=5)
    entry3.grid(row=5, column=0, padx=5, pady=5)

    # Увеличиваем ширину полей ввода
    entry1.config(width=70)
    entry2.config(width=70)
    entry3.config(width=70)

    # Создаем кнопку "Готово" и добавляем обработчик событий
    Button(root, text="Готово", command=lambda: handle_done_button(entry1, entry2, entry3)).grid(row=6, column=0, pady=10)

def renaming_of_professions():
    """Переименование профессий"""

    print("Переименование профессий") # Переименование профессий

    input_function_1()




if __name__ == '__main__':
    renaming_of_professions()
